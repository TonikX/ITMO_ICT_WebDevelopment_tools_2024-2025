import shutil
from datetime import datetime
import openpyxl
from ...config import TEMPLATE_EXCEL_PATH, FILLED_PROTOCOLS_DIR


def safe_write_cell(sheet, row, col, value):
    """
    Безопасно записывает значение в ячейку (row, col) с учётом объединённых диапазонов.
    row и col — это числовые индексы (R=row, C=col).
    """
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, max_col, max_row = merged_range.bounds
            top_left_cell = sheet.cell(row=min_row, column=min_col)
            top_left_cell.value = value
            return
    cell.value = value


def fill_excel_template(data):
    """
    Заполняет шаблон Excel данными из формы согласно указанным координатам.
    """
    if not TEMPLATE_EXCEL_PATH.exists():
        raise FileNotFoundError(f"Шаблон Excel не найден: {TEMPLATE_EXCEL_PATH}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filled_file_name = f"protocol_{timestamp}.xlsx"
    filled_file_path = FILLED_PROTOCOLS_DIR / filled_file_name

    # Копируем шаблон
    shutil.copy(TEMPLATE_EXCEL_PATH, filled_file_path)

    wb = openpyxl.load_workbook(filled_file_path)
    sheet = wb.active

    # -------------------------------
    # Хедер - базовая информация о матче
    # -------------------------------
    safe_write_cell(sheet, 2, 6, data.get('competition_type', ''))  # Вид соревнований
    safe_write_cell(sheet, 2, 17, data.get('match_date', ''))       # Дата
    safe_write_cell(sheet, 2, 22, data.get('match_number', ''))     # Игра №
    safe_write_cell(sheet, 3, 6, data.get('venue', ''))             # Место проведения
    safe_write_cell(sheet, 3, 17, data.get('match_time', ''))       # Время
    safe_write_cell(sheet, 3, 22, data.get('spectators', ''))       # Зрители

    # -------------------------------
    # Команда "A" (R6-R30)
    # -------------------------------
    # Название команды
    safe_write_cell(sheet, 5, 1, data.get('team_a', {}).get('name', ''))
    # Главный тренер
    safe_write_cell(sheet, 31, 8, data.get('team_a', {}).get('head_coach', ''))

    # Игроки команды А
    team_a_players = data.get('team_a', {}).get('players', [])
    for i, player in enumerate(team_a_players):
        if i >= 25:  # максимум 25 игроков (R6-R30)
            break
        row = 6 + i
        # Номер игрока
        safe_write_cell(sheet, row, 1, player.get('number', ''))
        # ФИО игрока
        safe_write_cell(sheet, row, 2, player.get('name', ''))
        # Капитан/Ассистент
        safe_write_cell(sheet, row, 8, player.get('captain', ''))
        # Позиция
        safe_write_cell(sheet, row, 9, player.get('position', ''))
        # Играл
        safe_write_cell(sheet, row, 10, 'ДА' if player.get('participated', False) else 'НЕТ')

    # Голы команды А
    goals_a = data.get('goals_a', [])
    for i, goal in enumerate(goals_a):
        if i >= 25:
            break
        row = 6 + i
        safe_write_cell(sheet, row, 11, goal.get('number', ''))     # №
        safe_write_cell(sheet, row, 12, goal.get('time', ''))       # Время
        safe_write_cell(sheet, row, 14, goal.get('scorer', ''))     # Гол
        safe_write_cell(sheet, row, 15, goal.get('assist1', ''))    # A
        safe_write_cell(sheet, row, 16, goal.get('assist2', ''))    # A
        safe_write_cell(sheet, row, 17, goal.get('situation', ''))  # ИС

    # Удаления команды А
    penalties_a = data.get('penalties_a', [])
    for i, penalty in enumerate(penalties_a):
        if i >= 25:
            break
        row = 6 + i
        safe_write_cell(sheet, row, 18, penalty.get('number', ''))      # №
        safe_write_cell(sheet, row, 19, penalty.get('time', ''))        # Штраф
        safe_write_cell(sheet, row, 20, penalty.get('infraction', ''))  # Причина
        safe_write_cell(sheet, row, 21, penalty.get('start', ''))       # Начало
        safe_write_cell(sheet, row, 23, penalty.get('end', ''))         # Окончание

    # -------------------------------
    # Команда "B" (R35-R59)
    # -------------------------------
    # Название команды
    safe_write_cell(sheet, 34, 1, data.get('team_b', {}).get('name', ''))
    # Главный тренер
    safe_write_cell(sheet, 60, 8, data.get('team_b', {}).get('head_coach', ''))

    # Игроки команды B
    team_b_players = data.get('team_b', {}).get('players', [])
    for i, player in enumerate(team_b_players):
        if i >= 25:  # максимум 25 игроков (R35-R59)
            break
        row = 35 + i
        # Номер игрока
        safe_write_cell(sheet, row, 1, player.get('number', ''))
        # ФИО игрока
        safe_write_cell(sheet, row, 2, player.get('name', ''))
        # Капитан/Ассистент
        safe_write_cell(sheet, row, 8, player.get('captain', ''))
        # Позиция
        safe_write_cell(sheet, row, 9, player.get('position', ''))
        # Играл
        safe_write_cell(sheet, row, 10, 'ДА' if player.get('participated', False) else 'НЕТ')

    # Голы команды B
    goals_b = data.get('goals_b', [])
    for i, goal in enumerate(goals_b):
        if i >= 25:
            break
        row = 35 + i
        safe_write_cell(sheet, row, 11, goal.get('number', ''))     # №
        safe_write_cell(sheet, row, 12, goal.get('time', ''))       # Время
        safe_write_cell(sheet, row, 14, goal.get('scorer', ''))     # Гол
        safe_write_cell(sheet, row, 15, goal.get('assist1', ''))    # A
        safe_write_cell(sheet, row, 16, goal.get('assist2', ''))    # A
        safe_write_cell(sheet, row, 17, goal.get('situation', ''))  # ИС

    # Удаления команды B
    penalties_b = data.get('penalties_b', [])
    for i, penalty in enumerate(penalties_b):
        if i >= 25:
            break
        row = 35 + i
        safe_write_cell(sheet, row, 18, penalty.get('number', ''))      # №
        safe_write_cell(sheet, row, 19, penalty.get('time', ''))        # Штраф
        safe_write_cell(sheet, row, 20, penalty.get('infraction', ''))  # Причина
        safe_write_cell(sheet, row, 21, penalty.get('start', ''))       # Начало
        safe_write_cell(sheet, row, 23, penalty.get('end', ''))         # Окончание

    # -------------------------------
    # Послематчевые броски (R64C1-R70C5)
    # -------------------------------
    shootout = data.get('shootout', [])
    for i, shot in enumerate(shootout):
        if i >= 7:  # Максимум 7 бросков (R64-R70)
            break
        row = 64 + i
        safe_write_cell(sheet, row, 1, shot.get('player_a', ''))    # А
        safe_write_cell(sheet, row, 2, shot.get('player_b', ''))    # Б
        safe_write_cell(sheet, row, 3, shot.get('goalie_a', ''))    # Вр А
        safe_write_cell(sheet, row, 4, shot.get('goalie_b', ''))    # Вр Б
        safe_write_cell(sheet, row, 5, shot.get('result', ''))      # Результат

    # -------------------------------
    # Время игры вратарей (R64C6-R70C9)
    # -------------------------------
    goalie_time = data.get('goalie_time', [])
    for i, gt in enumerate(goalie_time):
        if i >= 7:  # Максимум 7 записей (R64-R70)
            break
        row = 64 + i
        safe_write_cell(sheet, row, 6, gt.get('time', ''))         # Время
        safe_write_cell(sheet, row, 8, gt.get('goalie_a', ''))     # А
        safe_write_cell(sheet, row, 9, gt.get('goalie_b', ''))     # Б

    # -------------------------------
    # Результат по периодам (R63C14-R64C19)
    # -------------------------------
    periods = data.get('periods', {})
    # Первая строка - команда А
    safe_write_cell(sheet, 63, 14, periods.get('1', {}).get('a', ''))        # 1 период
    safe_write_cell(sheet, 63, 15, periods.get('2', {}).get('a', ''))        # 2 период
    safe_write_cell(sheet, 63, 16, periods.get('3', {}).get('a', ''))        # 3 период
    safe_write_cell(sheet, 63, 17, periods.get('ot', {}).get('a', ''))       # ОТ
    safe_write_cell(sheet, 63, 18, periods.get('so', {}).get('a', ''))       # ПБ
    safe_write_cell(sheet, 63, 19, periods.get('total', {}).get('a', ''))    # Общ

    # Вторая строка - команда Б
    safe_write_cell(sheet, 64, 14, periods.get('1', {}).get('b', ''))        # 1 период
    safe_write_cell(sheet, 64, 15, periods.get('2', {}).get('b', ''))        # 2 период
    safe_write_cell(sheet, 64, 16, periods.get('3', {}).get('b', ''))        # 3 период
    safe_write_cell(sheet, 64, 17, periods.get('ot', {}).get('b', ''))       # ОТ
    safe_write_cell(sheet, 64, 18, periods.get('so', {}).get('b', ''))       # ПБ
    safe_write_cell(sheet, 64, 19, periods.get('total', {}).get('b', ''))    # Общ

    # -------------------------------
    # Штрафное время (R65C14-R66C19)
    # -------------------------------
    penalties = data.get('penalties', {})

    # Команда А - первая строка
    safe_write_cell(sheet, 65, 14, penalties.get('1', {}).get('a', ''))  # 1 период
    safe_write_cell(sheet, 65, 15, penalties.get('2', {}).get('a', ''))  # 2 период
    safe_write_cell(sheet, 65, 16, penalties.get('3', {}).get('a', ''))  # 3 период
    safe_write_cell(sheet, 65, 17, penalties.get('ot', {}).get('a', ''))  # ОТ
    safe_write_cell(sheet, 65, 19, penalties.get('a', ''))  # Общ

    # Команда Б - вторая строка
    safe_write_cell(sheet, 66, 14, penalties.get('1', {}).get('b', ''))  # 1 период
    safe_write_cell(sheet, 66, 15, penalties.get('2', {}).get('b', ''))  # 2 период
    safe_write_cell(sheet, 66, 16, penalties.get('3', {}).get('b', ''))  # 3 период
    safe_write_cell(sheet, 66, 17, penalties.get('ot', {}).get('b', ''))  # ОТ
    safe_write_cell(sheet, 66, 19, penalties.get('b', ''))  # Общ

    # -------------------------------
    # Время матча и тайм-ауты
    # -------------------------------
    safe_write_cell(sheet, 63, 23, data.get('match_start_time', ''))    # Начало
    safe_write_cell(sheet, 64, 23, data.get('match_end_time', ''))      # Окончание

    timeouts = data.get('timeouts', {})
    safe_write_cell(sheet, 65, 23, timeouts.get('a', ''))              # Тайм-аут А
    safe_write_cell(sheet, 66, 23, timeouts.get('b', ''))              # Тайм-аут Б

    # -------------------------------
    # Судьи и официальные лица
    # -------------------------------
    officials = data.get('officials', {})
    # Судьи за воротами
    safe_write_cell(sheet, 67, 13, officials.get('goal_judge_1', ''))
    safe_write_cell(sheet, 68, 13, officials.get('goal_judge_2', ''))

    safe_write_cell(sheet, 69, 13, officials.get('announcer', ''))            # Информатор
    safe_write_cell(sheet, 70, 13, officials.get('timekeeper', ''))           # Судья времени игры

    # Судьи при оштрафованных
    safe_write_cell(sheet, 67, 20, officials.get('penalty_box_official_1', ''))
    safe_write_cell(sheet, 68, 20, officials.get('penalty_box_official_2', ''))

    safe_write_cell(sheet, 69, 20, officials.get('scorekeeper', ''))          # Секретарь

    # Главные судьи
    safe_write_cell(sheet, 71, 4, officials.get('main_referee', ''))
    safe_write_cell(sheet,71, 13, officials.get('reserve_referee', ''))       # Резервный судья

    # Сохраняем файл
    wb.save(filled_file_path)
    return str(filled_file_path)

